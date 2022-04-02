from ast import And, Not, Pass
from asyncio.windows_events import NULL
from cgitb import text
from distutils import command
from faulthandler import disable
from itertools import count
from optparse import Values
from pydoc import describe
from site import ENABLE_USER_SITE
from sqlite3 import Row
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from xmlrpc.client import boolean
from PIL import ImageTk, Image, ImageFile
import pandas as pd
from tkinter.messagebox import showinfo
import tkinter.scrolledtext as scrolledtext
from tkinter.filedialog import askopenfilename
import os
import webbrowser
from tkcalendar import Calendar
from tkcalendar import DateEntry
from datetime import date
from tkinter import filedialog
import subprocess
import mysql.connector
import io
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import shutil
import csv
fbilldb = mysql.connector.connect(
    host="localhost", user="root", password="", database="fbillingsintgrtd", port="3306"
)
fbcursor = fbilldb.cursor()

ImageFile.LOAD_TRUNCATED_IMAGES = True

def reset():
  global root
  root.destroy()

root=Tk()
root.geometry("1360x730")
root.resizable(False, False)
root.title("F-Billing Revolution 2022(FREE version) | Company database:fbillingdb | User:Administrator")
p1 = PhotoImage(file = 'images/fbicon.png')
root.iconphoto(False, p1)


s = ttk.Style()
s.theme_use('default')
s.configure('TNotebook.Tab', background="#999999", width=20, padding=10)
invoices= PhotoImage(file="images/invoice.png")
orders = PhotoImage(file="images/order.png")
estimates = PhotoImage(file="images/estimate.png")
recurring = PhotoImage(file="images/recurring.png")
purchase = PhotoImage(file="images/purchase.png")
expenses = PhotoImage(file="images/expense.png")
customer = PhotoImage(file="images/customer.png")
product = PhotoImage(file="images/package.png")
reports = PhotoImage(file="images/report.png")
setting = PhotoImage(file="images/setting.png")
tick = PhotoImage(file="images/check.png")
warnin = PhotoImage(file="images/sign_warning.png")
cancel = PhotoImage(file="images/close.png")
saves = PhotoImage(file="images/save.png")
folder = PhotoImage(file="images/folder-black.png")
photo11 = PhotoImage(file = "images/invoice-pvt.png")
customer = PhotoImage(file="images/customer.png")
smslog = PhotoImage(file = "images/smslog.png")
video = PhotoImage(file = "images/video.png")
mark1 = PhotoImage(file="images/mark.png")
mark2 = PhotoImage(file="images/mark2.png")
photo10 = PhotoImage(file = "images/text-message.png")
addnew = PhotoImage(file="images/plus.png")
delete = PhotoImage(file="images/delete_E.png")
tabControl = ttk.Notebook(root)
tab1 = ttk.Frame(tabControl)
tab2 = ttk.Frame(tabControl)
tab3=  ttk.Frame(tabControl)
tab4 = ttk.Frame(tabControl)
tab5 = ttk.Frame(tabControl)
tab6=  ttk.Frame(tabControl)
tab7 = ttk.Frame(tabControl)
tab8 = ttk.Frame(tabControl)
tab9 =  ttk.Frame(tabControl)
tab10=  ttk.Frame(tabControl)
tabControl.add(tab1,image=invoices,compound = LEFT, text ='Invoices',)
tabControl.add(tab2,image=orders,compound = LEFT, text ='Orders')
tabControl.add(tab3,image=estimates,compound = LEFT, text ='Estimates')
tabControl.add(tab4,image=recurring,compound = LEFT, text ='Recurring')
tabControl.add(tab5,image=purchase,compound = LEFT, text ='Purchase Orders') 
tabControl.add(tab6,image=expenses,compound = LEFT, text ='Expenses')
tabControl.add(tab7,image=customer,compound = LEFT, text ='Customers')
tabControl.add(tab8,image=product,compound = LEFT, text ='Product/Services')
tabControl.add(tab9,image=reports,compound = LEFT, text ='Report')
tabControl.add(tab10,image=setting,compound = LEFT, text ='Settings')
tabControl.pack(expand = 1, fill ="both")


selectall = PhotoImage(file="images/table_select_all.png")
cut = PhotoImage(file="images/cut.png")
copy = PhotoImage(file="images/copy.png")
paste = PhotoImage(file="images/paste.png")

undo = PhotoImage(file="images/undo.png")
redo = PhotoImage(file="images/redo.png")
bold = PhotoImage(file="images/bold.png")

italics = PhotoImage(file="images/italics.png")
underline = PhotoImage(file="images/underline.png")
left = PhotoImage(file="images/left.png")

right = PhotoImage(file="images/right.png")
center = PhotoImage(file="images/center.png")
hyperlink = PhotoImage(file="images/hyperlink.png")
remove = PhotoImage(file="images/eraser.png")


photo = PhotoImage(file = "images/plus.png")
photo1 = PhotoImage(file = "images/edit.png")
photo2 = PhotoImage(file = "images/delete_E.png")
photo3 = PhotoImage(file = "images/export-file.png")
photo4 = PhotoImage(file = "images/seo.png")
photo5 = PhotoImage(file = "images/printer.png")
photo6 = PhotoImage(file = "images/gmail.png")
photo7 = PhotoImage(file = "images/priewok.png")
photo8 = PhotoImage(file = "images/refresh_E.png")
photo9 = PhotoImage(file = "images/sum.png")
photo10 = PhotoImage(file = "images/text-message.png")


def check_empty() :
     if entry.get():
         pass     #your function where you want to jump
     else:
        messagebox.showinfo("Information", "Required entry")

def adda_product():
    top = Toplevel()  
    
    top.title("Add a new Product/Service")
    p2 = PhotoImage(file = 'images/fbicon.png')
    top.iconphoto(False, p1)
    top.geometry("600x550+390+125")
    
    
    tabControl = ttk.Notebook(top)
    s = ttk.Style()
    s.theme_use('default')
    s.configure('TNotebook.Tab', background="#999999", width=50, padding=10,bd=0)


    tab1 = ttk.Frame(tabControl)
    tab2 = ttk.Frame(tabControl)
    
    tabControl.add(tab1,compound = LEFT, text ='Product/Service')
    tabControl.add(tab2,compound = LEFT, text ='Product Image')
    
    tabControl.pack(expand = 1, fill ="both")
    
    innerFrame = Frame(tab1,bg="#f5f3f2", relief=GROOVE, height=490)
    innerFrame.pack(side="top",fill=BOTH)

    Customerlabelframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=475)
    Customerlabelframe.pack(side="top",fill=BOTH,padx=10)
    
    global filename
    filename = ""
    
    def upload_file():
      global filename,img, b2
      f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
      filename = filedialog.askopenfilename(filetypes=f_types)
      shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
      image = Image.open(filename)
      resize_image = image.resize((350, 350))
      img = ImageTk.PhotoImage(resize_image)
      b2 = Button(imageFrame,image=img)
      b2.place(x=130, y=80)
    
    def addproducts():
      global img , filename 
      sku = codeentry.get()
      status = checkvarStatus.get()
      catgory = n.get()
      name = nameentry.get()
      description = desentry.get()
      unitprice = uval.get()
      peices = pcsentry.get()
      cost = costval.get()
      price_cost = priceval.get()
      taxable = checkvarStatus2.get()
      nostockcontrol = checkvarStatus3.get()
      stock = stockval.get()
      lowstock = lowval.get()
      warehouse = wareentry.get()
      pnotes = sctxt.get("1.0",'end-1c')
      entries = [sku,name, unitprice, cost]
      entri = []
      for i in entries:
        if i == '':
          entri.append(i)
      if len(entri) == 0:
        sql = 'select * from Productservice where sku = %s or name = %s'
        val  = (sku, name)
        fbcursor.execute(sql, val)
        fbcursor.fetchall()
        row_count = fbcursor.rowcount
        if row_count == 0:
          if filename == "":
            sql = 'insert into Productservice(sku, category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
            val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes)
            fbcursor.execute(sql, val)
            fbilldb.commit()
          else:
            file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            sql = 'insert into Productservice(sku, category, name, description, status, unitprice, peices, cost, taxable, priceminuscost, serviceornot, stock, stocklimit, warehouse, image, privatenote) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'
            val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, filename.split('/')[-1], pnotes)
            fbcursor.execute(sql, val)
            fbilldb.commit()
        else:
          messagebox.showinfo("Alert", "Entry with same name or SKU already exists.\nTry again.")
        top.destroy()
        messagebox.showinfo("F-Billing Revolution", "Product added successfully.")
        for record in treeproducts.get_children():
          treeproducts.delete(record)
        fbcursor.execute("select *  from Productservice")
        pandsdata = fbcursor.fetchall()
        countp = 0
        for i in pandsdata:
          if i[6] == '1':
            acti = 'Active'
          else:
            acti = 'Inactive' 
          if i[13] > i[14]:
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
            countp += 1
          elif (i[12] =="0") == (i[13] <= i[14]):
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
            countp += 1
          elif i[12] == '1':
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
            countp += 1
          else:
            pass
          
       
      else:
        messagebox.showinfo("Alert", "Fields name and SKU should not be empty.\nFill out required fields and try again")
        top.destroy()
 
    fbcursor.execute("SELECT * FROM Productservice ORDER BY sku DESC LIMIT 1")
    skuin = fbcursor.fetchone()
    
    
    code1=Label(Customerlabelframe,text="Code or SKU* :",fg="blue",pady=10,padx=10)
    code1.place(x=20,y=0)
    codeentry = Entry(Customerlabelframe,width=35)
    codeentry.place(x=110,y=8)
    # if not skuin == None:
    #   fk=skuin[2]+1
    # else:
    #   fk=1
    # codeentry.insert(0, fk)

    checkvarStatus=IntVar()
    status1=Label(Customerlabelframe,text="Status:")
    status1.place(x=380,y=8)
    Button1 = Checkbutton(Customerlabelframe, 
                      variable = checkvarStatus,text="Active",compound="right",
                      onvalue =1,
                      offvalue = 0,
                      width = 10)

    Button1.place(x=420,y=5)

    category1=Label(Customerlabelframe,text="Category:",pady=5,padx=10)
    category1.place(x=20,y=40)
    n = StringVar() 
    catgory = ttk.Combobox(Customerlabelframe, width = 40, textvariable = n ) 
    catgory.place(x=110,y=45)
    catgory.insert(0, 'Default')


    name1=Label(Customerlabelframe,text="Name* :",fg="blue",pady=5,padx=10)
    name1.place(x=20,y=70)
    nameentry = Entry(Customerlabelframe,width=70)
    nameentry.place(x=110,y=75)

    des1=Label(Customerlabelframe,text="Description :",pady=5,padx=10)
    des1.place(x=20,y=100)
    desentry = Entry(Customerlabelframe,width=70)
    desentry.place(x=110,y=105)

    uval = IntVar()
    unit1=Label(Customerlabelframe,text="Unit Price:",fg="blue",pady=5,padx=10)
    unit1.place(x=20,y=130)
    unitentry = Entry(Customerlabelframe,width=20,textvariable=uval)
    unitentry.place(x=110,y=135)
    

    # pcsval = IntVar()
    pcs1=Label(Customerlabelframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
    pcs1.place(x=320,y=130)
    pcsentry = Entry(Customerlabelframe,width=20)
    pcsentry.place(x=410,y=135)

    costval = IntVar()
    cost1=Label(Customerlabelframe,text="Cost:",pady=5,padx=10)
    cost1.place(x=20,y=160)
    
    
    costentry = Entry(Customerlabelframe,width=20,textvariable=costval)
    costentry.place(x=110,y=165)
    # unit = int(uval.get())
    # cost = int(costentry.get())
    # prccst = unit-cost
    def set_label(name, index, mode):
      priceval.set(uval.get() - costval.get())
    priceval = IntVar()
    price1=Label(Customerlabelframe,text="(Price-Cost):",pady=5,padx=10)
    price1.place(x=20,y=190)
    priceentry = Entry(Customerlabelframe,width=20,textvariable=priceval,state=DISABLED,disabledbackground="white",disabledforeground="black")
    priceentry.place(x=110,y=195)
    
    uval.trace('w', set_label)
    costval.trace('w', set_label)

    checkvarStatus2=IntVar()
   
    Button2 = Checkbutton(Customerlabelframe,variable = checkvarStatus2, 
                      text="Taxable Tax1rate",compound="right",
                      onvalue =1 ,
                      offvalue = 0,
                      height=2,
                      width = 12)

    Button2.place(x=415,y=153)

    def switch():
      if checkvarStatus3.get():
        stockentry["state"] = DISABLED
        lowentry["state"] = DISABLED
        wareentry["state"] = DISABLED
      else:
        stockentry["state"] = NORMAL
        lowentry["state"] = NORMAL
        wareentry["state"] = NORMAL
    checkvarStatus3=BooleanVar()
    Button3 = Checkbutton(Customerlabelframe,variable = checkvarStatus3,command=switch, 
                      text="This is a service(no stock control)", 
                      onvalue =1 ,
                      offvalue = 0,
                      height=3)

    Button3.place(x=40,y=220)


    stockval = IntVar()
    stock1=Label(Customerlabelframe,text="Stock:",pady=5,padx=10)
    stock1.place(x=90,y=260)
    stockentry = Entry(Customerlabelframe,width=15,textvariable=stockval)
    stockentry.place(x=140,y=265)

    lowval = IntVar()
    low1=Label(Customerlabelframe,text="Low Stock Warning Limit:",pady=5,padx=10)
    low1.place(x=280,y=260)
    lowentry = Entry(Customerlabelframe,width=15,textvariable=lowval)
    lowentry.place(x=435,y=265)

   
    ware1=Label(Customerlabelframe,text="Warehouse:",pady=5,padx=10)
    ware1.place(x=60,y=290)
    wareentry = Entry(Customerlabelframe,width=64)
    wareentry.place(x=140,y=295)

    # pnoteval = StringVar()
    text1=Label(Customerlabelframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
    text1.place(x=20,y=320)
    sctxt = scrolledtext.ScrolledText(Customerlabelframe, undo=True,width=62,height=4)
    sctxt.place(x=32,y=358)
    
    okButton = Button(innerFrame, text ="Ok",image=tick,width=70,compound = LEFT, command=addproducts)
    okButton.pack(side=LEFT, padx=(10, 0), pady=(5, 10))
    
    def closetab():
      top.destroy()

    cancelButton = Button(innerFrame,image=cancel,text="Cancel",width=70,compound = LEFT, command=closetab)
    cancelButton.pack(side=RIGHT, padx=(0, 10), pady=(5, 10))

    imageFrame = Frame(tab2, relief=GROOVE,height=580)
    imageFrame.pack(side="top",fill=BOTH)

    
      
    browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
    browseimg.place(x=30,y=35)
      
    browsebutton=Button(imageFrame,text = 'Browse',command=upload_file)
    browsebutton.place(x=485,y=30,height=30,width=50)

    removeButton = Button(imageFrame,image=cancel,text="Remove Product Image",width=150,compound = LEFT, command=lambda: b2.destroy())
    removeButton.place(x=410,y=460)

    top.mainloop()

def fileimport_product():

    top=Toplevel()
    top.title("Import items list from Excel(XLS)File")
    top.geometry("785x520+280+100")
    importframe=Frame(top)
    importframe.place(x=0,y=0,height=700,width=785)
    impolbl=Label(importframe,text="Import source Excel(xlsx) File:").place(x=8,y=30)
    impoentry=Entry(importframe,bg="white")
    impoentry.place(x=8,y=50,width=280, height=25)
    previewlbl=Label(importframe,text="Source File preview").place(x=8,y=77)
   
    ###### LISTBOX #####################
    scrollbarx = Scrollbar(importframe, orient=HORIZONTAL)
    scrollbary = Scrollbar(importframe, orient=VERTICAL)
    imptree = ttk.Treeview(importframe, columns=("PRODUCT SERVICE ID","CODE OR SKU","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"), height=400,     selectmode="extended", yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
    scrollbary.config(command=imptree.yview)
    scrollbary.place(x=354,y=100,height=325)
    scrollbarx.config(command=imptree.xview)
    scrollbarx.place(x=0,y=410, width=356)
    imptree.heading('PRODUCT SERVICE ID', text="PRODUCT SERVICE ID", anchor=W)
    imptree.heading('CODE OR SKU', text="CODE OR SKU", anchor=W)
    imptree.heading('NAME', text="NAME", anchor=W)
    imptree.heading('CATEGORY', text="CATEGORY", anchor=W)
    imptree.heading('NAME', text="NAME", anchor=W)
    imptree.heading('DESCRIPTION', text="DESCRIPTION", anchor=W)
    imptree.heading('QTY UNIT', text="QTY UNIT", anchor=W)
    imptree.heading('COST', text="COST", anchor=W)
    imptree.heading('PRICE', text="PRICE", anchor=W)
    imptree.heading('TAX1', text="TAX1", anchor=W)
    imptree.heading('STOCK', text="STOCK", anchor=W)
    imptree.heading('LOW STOCK', text="LOW STOCK", anchor=W)
    imptree.heading('LOCATION', text="LOCATION", anchor=W)
    imptree.heading('ACTIVE', text="ACTIVE", anchor=W)
    imptree.heading('SERVICE', text="SERVICE", anchor=W)
    

    imptree.column('#0', stretch=NO, minwidth=0, width=0)
    imptree.column('#1', stretch=NO, minwidth=0, width=120)
    imptree.column('#2', stretch=NO, minwidth=0, width=100)
    imptree.column('#3', stretch=NO, minwidth=0, width=100)
    imptree.column('#4', stretch=NO, minwidth=0, width=100)
    imptree.column('#5', stretch=NO, minwidth=0, width=100)
    imptree.column('#6', stretch=NO, minwidth=0, width=100)
    imptree.column('#7', stretch=NO, minwidth=0, width=100)
    imptree.column('#8', stretch=NO, minwidth=0, width=100)
    imptree.column('#9', stretch=NO, minwidth=0, width=100)
    imptree.column('#10', stretch=NO, minwidth=0, width=100)
    imptree.column('#12', stretch=NO, minwidth=0, width=100)
    imptree.column('#13', stretch=NO, minwidth=0, width=100)
    imptree.column('#14', stretch=NO, minwidth=0, width=100)

 

    imptree.place(x=0,y=100,height=315,width=356)
    # langs = ()

    # langs_var = StringVar(value=langs)

    # listbox = Listbox(
    #     importframe,
    #     listvariable=langs_var,
    #     width=60,
    #     height=6,
    #     selectmode='extended')

    # listbox.place(x=8,y=100,height=320)
    
    # link a scrollbar to a list
    # scrollbar = Scrollbar(
    #     importframe,
    #     orient='vertical',
    #     command=listbox.yview
    # )
    
    # listbox['yscrollcommand'] = scrollbar.set
    # scrollbar.place(x=354,y=100,height=319)


    # scrollbar12 = Scrollbar(
    #     importframe,
    #     orient='horizontal',
    #     command=listbox.xview 
    # )
    # listbox['xscrollcommand'] = scrollbar.set
    # scrollbar12.place(x=0,y=402, width=354)

    
    

    lb1=Label(importframe,text="Select import source XLs file first after build column associations").place(x=8,y=480)

    

    def export_product_1():
      global Productserviceid,name12,category12,description,peices,cost12,priceminuscost,taxable,stock12,stocklimit,warehouse,status,serviceornot,name
      name = askopenfilename(filetypes=[('CSV', '*.csv',), ('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
      # df = pd.read_csv(name)
      # for i in df:
      #   listbox.insert(END, df)
      with open(name) as f:
        reader = csv.DictReader(f, delimiter=',')
        print(reader)
        for row in reader:
          # "PRODUCT SERVICE ID","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"
          Productserviceid = row['PRODUCT SERVICE ID']
          sku = row['CODE OR SKU']          
          name12 = row['NAME']
          category12 = row['CATEGORY']
          
          description = row['DESCRIPTION']
          peices = row['QTY UNIT']
          cost12 = row['COST']
          priceminuscost = row['PRICE']
          taxable = row['TAX1']
          stock12 = row['STOCK']
          stocklimit = row['LOW STOCK']
          warehouse = row['LOCATION']
          status = row['ACTIVE']
          serviceornot = row['SERVICE']
        
          imptree.insert("", 0, values=(Productserviceid,sku,name12,category12,description,peices,cost12,priceminuscost,taxable,stock12,stocklimit,warehouse,status,serviceornot))
          

        
          # print(row)
          # listbox.insert(END, row)
      impoentry.delete(0, 'end')
      impoentry.insert(0, name)
      
    def nxtscreen():
      def save_pro_import():
        with open(name) as f:
          reader = csv.DictReader(f, delimiter=',')
          for row in reader:
          # "PRODUCT SERVICE ID","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"
            Productserviceid = int(row['PRODUCT SERVICE ID'])
            sku = int(row['CODE OR SKU'])  
            name12 = row['NAME']
            category12 = row['CATEGORY']
            description = row['DESCRIPTION']
            peices = int(row['QTY UNIT'])
            cost12 = int(row['COST'])
            unitprice = int(row['PRICE'])
            taxable = int(row['TAX1'])
            stock12 = int(row['STOCK'])
            stocklimit = int(row['LOW STOCK'])
            warehouse = row['LOCATION']
            status = int(row['ACTIVE'])
            serviceornot = int(row['SERVICE'])
            min = int(unitprice) - int(cost12)

            sql = 'select * from Productservice where Productserviceid = %s or name = %s or sku=%s'
            val  = (Productserviceid, name12,sku)
            fbcursor.execute(sql, val)
            fbcursor.fetchall()
            row_count = fbcursor.rowcount
            if row_count == 0:
              sql = 'insert into Productservice(Productserviceid,sku,name,category,description,peices,cost,unitprice,taxable,stock,stocklimit,warehouse,status,serviceornot,priceminuscost) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s)'
              val = (Productserviceid,sku,name12,category12,description,peices,cost12,unitprice,taxable,stock12,stocklimit,warehouse,status,serviceornot,min)
              fbcursor.execute(sql, val)
              fbilldb.commit()
              topp.destroy()
              for record in treeproducts.get_children():
                treeproducts.delete(record)
              fbcursor.execute("select *  from Productservice")
              pandsdata = fbcursor.fetchall()
              countp = 0
              for i in pandsdata:
                if i[6] == '1':
                  acti = 'Active'
                else:
                  acti = 'Inactive' 
                if i[13] > i[14]:
                  treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3],acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
                  countp += 1
                elif (i[12] =="0") == (i[13] <= i[14]):
                  treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3],acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
                  countp += 1
                elif i[12] == '1':
                  treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3],acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
                  countp += 1
                else:
                  pass
            else:
              messagebox.showinfo("Alert", "Entry with same name or SKU already exists.\nTry again.")
            
            

            # if len(entri) == 0:
            #   sql = 'select * from Productservice where sku = %s or name = %s'
            #   val  = (sku, name)
            #   fbcursor.execute(sql, val)
            #   fbcursor.fetchall()
            #   row_count = fbcursor.rowcount
            #   if row_count == 0:
            #     if filename == "":
                  
      
        

     
      topp=Toplevel()
      topp.title("Import items list from Excel(XLS)File")
      topp.geometry("785x520+280+100")
      scrollbarx = Scrollbar(topp, orient=HORIZONTAL)
      scrollbary = Scrollbar(topp, orient=VERTICAL)
      nxttree = ttk.Treeview(topp, columns=("PRODUCT SERVICE ID","CODE OR SKU","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"),height=400,     selectmode="extended", yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
      scrollbary.config(command=nxttree.yview)
      scrollbary.place(x=768,y=0,height=490)
      scrollbarx.config(command=nxttree.xview)
      scrollbarx.place(x=0,y=470,width=763)
      nxttree.heading('CODE OR SKU', text="CODE OR SKU", anchor=W)
      nxttree.heading('PRODUCT SERVICE ID', text="PRODUCT SERVICE ID", anchor=W)
      nxttree.heading('NAME', text="NAME", anchor=W)
      nxttree.heading('CATEGORY', text="CATEGORY", anchor=W)
      nxttree.heading('NAME', text="NAME", anchor=W)
      nxttree.heading('DESCRIPTION', text="DESCRIPTION", anchor=W)
      nxttree.heading('QTY UNIT', text="QTY UNIT", anchor=W)
      nxttree.heading('COST', text="COST", anchor=W)
      nxttree.heading('PRICE', text="PRICE", anchor=W)
      nxttree.heading('TAX1', text="TAX1", anchor=W)
      nxttree.heading('STOCK', text="STOCK", anchor=W)
      nxttree.heading('LOW STOCK', text="LOW STOCK", anchor=W)
      nxttree.heading('LOCATION', text="LOCATION", anchor=W)
      nxttree.heading('ACTIVE', text="ACTIVE", anchor=W)
      nxttree.heading('SERVICE', text="SERVICE", anchor=W)
  
      nxttree.column('#0', stretch=NO, minwidth=0, width=0)
      nxttree.column('#1', stretch=NO, minwidth=0, width=120)
      nxttree.column('#2', stretch=NO, minwidth=0, width=100)
      nxttree.column('#3', stretch=NO, minwidth=0, width=100)
      nxttree.column('#4', stretch=NO, minwidth=0, width=100)
      nxttree.column('#5', stretch=NO, minwidth=0, width=100)
      nxttree.column('#6', stretch=NO, minwidth=0, width=100)
      nxttree.column('#7', stretch=NO, minwidth=0, width=100)
      nxttree.column('#8', stretch=NO, minwidth=0, width=100)
      nxttree.column('#9', stretch=NO, minwidth=0, width=100)
      nxttree.column('#10', stretch=NO, minwidth=0, width=100)
      nxttree.column('#12', stretch=NO, minwidth=0, width=100)
      nxttree.column('#13', stretch=NO, minwidth=0, width=100)
      nxttree.column('#14', stretch=NO, minwidth=0, width=100)
    
      with open(name) as f:
        reader = csv.DictReader(f, delimiter=',')
        for row in reader:
          # "PRODUCT SERVICE ID","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"
          Productserviceid = row['PRODUCT SERVICE ID']
          sku = row['CODE OR SKU']
          name12 = row['NAME']
          category12 = row['CATEGORY']
          
          description = row['DESCRIPTION']
          peices = row['QTY UNIT']
          cost12 = row['COST']
          priceminuscost = row['PRICE']
          taxable = row['TAX1']
          stock12 = row['STOCK']
          stocklimit = row['LOW STOCK']
          warehouse = row['LOCATION']
          status = row['ACTIVE']
          serviceornot = row['SERVICE']
      

      
          nxttree.insert("", 0, values=(Productserviceid,sku,name12,category12,description,peices,cost12,  priceminuscost,taxable,stock12,stocklimit,warehouse,status,serviceornot))
       
    
      nxttree.place(x=0,y=0,height=470,width=770)
      back = Button(topp,text="back",command=lambda:topp.destroy())
      back.place(x=5,y=492)
      Finish = Button(topp,text="Finish",command=save_pro_import)
      Finish.place(x=740,y=492)
     
      


      

      # name = askopenfilename(filetypes=[('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
      # impoentry.delete(0, 'end')
      # impoentry.insert(0, name)
      # wb = Workbook()
      # wb = load_workbook(name)
      # ws = wb.active
      # column_a = ws['A']
      # column_b = ws['B']
      # paslist = []
      # for cell in column_b:
      #     paslist.append(str(cell.value))
      
      # pslist = paslist
      # nameentry.insert(0, pslist[0])
      # category.insert(0, pslist[1])
      # desc.insert(0, pslist[2])
      # quantity.insert(0, pslist[3])
      # cost.insert(0, pslist[4])
      # price.insert(0, pslist[5])
      # tax.insert(0, pslist[6])
      # tax2.insert(0, pslist[7])
      # stock2.insert(0, pslist[8])
      # lowstock.insert(0, pslist[9])
      # location.insert(0, pslist[10])
      # active.insert(0, pslist[11])
      # service.insert(0, pslist[12])
      
      # paslist.reverse()
      # listbox.delete(0, 'end')
      # for i in paslist:
      #   listbox.insert(0, f'{i}\n')
      
    
    importbutton=Button(top,command=export_product_1,text = 'Browse',compound=LEFT)
    importbutton.place(x=290,y=48,height=25,width=80)

    lb1=Label(importframe,text="Please associate datafields with data columns").place(x=460,y=30)

    id=Label(importframe,text="PRODUCT SERVICE ID = ",pady=5,padx=10,fg="blue")
    id.place(x=430,y=90)
    idd = StringVar() 
    idds = ttk.Combobox(importframe, width = 30, textvariable = idd) 
    idds.place(x=570,y=96) 

    name1=Label(importframe,text="NAME = ",pady=5,padx=10,fg="blue")
    name1.place(x=450,y=115)
    namevar = StringVar() 
    nameentry = ttk.Combobox(importframe, width = 30, textvariable = namevar ) 
    nameentry.place(x=570,y=121)

   
    category1=Label(importframe,text="CATEGORY = ",pady=5,padx=10,fg="blue")
    category1.place(x=450,y=140)
    categoryvar = StringVar() 
    category = ttk.Combobox(importframe, width = 30, textvariable = categoryvar ) 
    category.place(x=570,y=146)

    desc1=Label(importframe,text="DESCRIPTION = ",pady=5,padx=10)
    desc1.place(x=450,y=165)
    
    descvar = StringVar() 
    desc = ttk.Combobox(importframe, width = 30, textvariable = descvar ) 
    desc.place(x=570,y=171)

    quan1=Label(importframe,text="QUANTITY= ",pady=5,padx=10)
    quan1.place(x=450,y= 190)
    
    quanvar = StringVar() 
    quantity = ttk.Combobox(importframe, width = 30, textvariable = categoryvar ) 
    quantity.place(x=570,y=196)

    cost1=Label(importframe,text="COST = ",pady=5,padx=10)
    cost1.place(x=450,y=215)
    
    costvar = StringVar() 
    cost = ttk.Combobox(importframe, width = 30, textvariable = costvar ) 
    cost.place(x=570,y=221)

    price1=Label(importframe,text="PRICE = ",pady=5,padx=10,fg="blue")
    price1.place(x=450,y=240)
    
    pricevar = StringVar() 
    price = ttk.Combobox(importframe, width = 30, textvariable = costvar ) 
    price.place(x=570,y=246)


    taxx1=Label(importframe,text="TAX1 = ",pady=5,padx=10,)
    taxx1.place(x=450,y=265)
    taxvar = StringVar() 
    tax = ttk.Combobox(importframe, width = 30, textvariable = taxvar ) 
    tax.place(x=570,y=271)

    # taxx2=Label(importframe,text="TAX2 = ",pady=5,padx=10)
    # taxx2.place(x=450,y=265)
    # tax2var = StringVar() 
    # tax2 = Entry(importframe, width = 30, textvariable = tax2var ) 
    # tax2.place(x=570,y=271)

    stock2=Label(importframe,text="STOCK = ",pady=5,padx=10)
    stock2.place(x=450,y=290)
    
    stock2var = StringVar() 
    stock2 = ttk.Combobox(importframe, width = 30, textvariable = stock2var ) 
    stock2.place(x=570,y=296)


    lowstock2=Label(importframe,text="LOW STOCK = ",pady=5,padx=10)
    lowstock2.place(x=450,y=315)
    
    lowstock2var = StringVar() 
    lowstock = ttk.Combobox(importframe, width = 30, textvariable = lowstock2var ) 
    lowstock.place(x=570,y=321)


    location2=Label(importframe,text="LOCATION = ",pady=5,padx=10)
    location2.place(x=450,y=340)
    Location2var = StringVar() 
    location = ttk.Combobox(importframe, width = 30, textvariable = Location2var ) 
    location.place(x=570,y=346)

    active2=Label(importframe,text="ACTIVE = ",pady=5,padx=10)
    active2.place(x=450,y=365)
    
    active2var = StringVar() 
    active = ttk.Combobox(importframe, width = 30, textvariable = active2var ) 
    active.place(x=570,y=371)

    service2=Label(importframe,text="SERVICE = ",pady=5,padx=10)
    service2.place(x=450,y=390)
    
    service2var = StringVar() 
    service = ttk.Combobox(importframe, width = 30, textvariable = service2var ) 
    service.place(x=570,y=396)

    b = Button(importframe,text = "Clear associations").place(x=600,y=470)  
    n = Button(importframe, text ="Next",command=nxtscreen).place(x=710,y=470)
  
    
    top.mainloop()
    
#########EXPORT PRODUCT#######################################################################################

def export_product():
  cols = ["PRODUCT SERVICE ID","CODE OR SKU","NAME","CATEGORY","DESCRIPTION","QTY UNIT","COST","PRICE","TAX1","STOCK","LOW STOCK","LOCATION","ACTIVE","SERVICE"] # Your column headings here
  path = filedialog.asksaveasfilename(initialdir=os.getcwd,title="Save File",filetypes=[('CSV', '*.csv',), ('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
  excel_name = 'newfile.xlsx'
  lst = []
  with open(path, "w", newline='') as myfile:
      csvwriter = csv.writer(myfile, delimiter=',')
      sql = 'select Productserviceid,sku,name,category,description,peices,cost,unitprice,taxable,stock,stocklimit,warehouse,status,serviceornot from Productservice'
      fbcursor.execute(sql)
      pandsdata = fbcursor.fetchall()
      print (pandsdata)
      for row_id in pandsdata:
          row = row_id
          lst.append(row)
      lst = list(map(list,lst))
      lst.insert(0,cols)
      for row in lst:
          csvwriter.writerow(row)

  writer = pd.ExcelWriter(excel_name)
  df = pd.read_csv(path)
  df.to_excel(writer,'sheetname')
  writer.save()
    # name = filedialog.asksaveasfilename(initialdir=os.getcwd,title="Save File",filetypes=[('CSV', '*.csv',), ('Excel', ('*.xls', '*.xslm', '*.xlsx'))])
    # with open(name,mode='w',newline='') as myfile:
    #   exp_writer = csv.writer(myfile,delimiter='\t')
    #   for row_id in treeproducts.get_children():
    #       row = treeproducts.item(row_id)['values']
    #       exp_writer.writerow(row)
        
    #   # sql = 'select * from Productservice'
    #   # fbcursor.execute(sql)
    #   # pandsdata = fbcursor.fetchall()
    #   # for i in pandsdata:
    #   #   exp_writer.writerow(i)
    # messagebox.showinfo("saved")

######################## VIEW/EDIT PRODUCT #######################################################################

  
def edit_product():  
    try:
      itemid = treeproducts.item(treeproducts.focus())["values"][1]
      
      global filename
      filename = ""
      
      def upload_file():
        global filename,img, b2
        f_types =[('Png files','*.png'),('Jpg Files', '*.jpg')]
        filename = filedialog.askopenfilename(filetypes=f_types)
        shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
        image = Image.open(filename)
        resize_image = image.resize((350, 350))
        img = ImageTk.PhotoImage(resize_image)
        b2 = Button(imageFrame,image=img)
        b2.place(x=130, y=80)
      
      def updateproducts():
        global img , filename 
        sku = codeentry.get()
        status = checkvarStatus.get()
        catgory = n.get()
        name = nameentry.get()
        description = desentry.get()
        unitprice = uval.get()
        peices = pcsentry.get()
        cost = costval.get()
        price_cost = priceval.get()
        taxable = checkvarStatus2.get()
        nostockcontrol = checkvarStatus3.get()
        stock = stockval.get()
        lowstock = lowval.get()
        warehouse = wareentry.get()
        pnotes = sctxt.get("1.0", 'end-1c')
        entries = [sku, name, unitprice, cost]
        entri = []
        for i in entries:
          if i == '':
            entri.append(i)
        if len(entri) == 0:
          if filename == "":
            sql = "update Productservice set sku=%s, category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, privatenote=%s where Productserviceid = %s"
            val = (sku, catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse, pnotes, itemid)
            fbcursor.execute(sql, val)
            fbilldb.commit()
          else:
            file = shutil.copyfile(filename, os.getcwd()+'/images/'+filename.split('/')[-1])
            sql = "update Productservice set category=%s, name=%s, description=%s, status=%s, unitprice=%s, peices=%s, cost=%s, taxable=%s, priceminuscost=%s, serviceornot=%s, stock=%s, stocklimit=%s, warehouse=%s, image=%s, privatenote=%s where Productserviceid = %s"
            val = (catgory, name, description, status, unitprice, peices, cost, taxable, price_cost, nostockcontrol, stock, lowstock, warehouse,filename.split('/')[-1], pnotes, itemid)
            fbcursor.execute(sql, val)
            fbilldb.commit()
            messagebox.showinfo("F-Billing Revolution", "Product updated successfully.")
            for record in treeproducts.get_children():
              treeproducts.delete(record)
            fbcursor.execute("select *  from Productservice")
            pandsdata = fbcursor.fetchall()
            countp = 0
            for i in pandsdata:
              if i[6] == '1':
                acti = 'Active'
              else:
                acti = 'Inactive' 
              if i[13] > i[14]:
                treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('green',))
                countp += 1
              elif (i[12] =="0") == (i[13] <= i[14]):
                treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4]  , i[7], i[13], i[15],i[17]),tags=('red',))
                countp += 1
              elif i[12] == '1':
                treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('blue',))
                countp += 1
              else:
                pass
          top.destroy()
        else:
          messagebox.showinfo("F-Billing Revolution", "Fields name or SKU entered is already in database.")
          top.destroy()
        for record in treeproducts.get_children():
              treeproducts.delete(record)
        fbcursor.execute("select *  from Productservice")
        pandsdata = fbcursor.fetchall()
        countp = 0
        for i in pandsdata:
          if i[6] == '1':
            acti = 'Active'
          else:
            acti = 'Inactive' 
          if i[13] > i[14]:
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('green',))
            countp += 1
          elif (i[12] =="0") == (i[13] <= i[14]):
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('red',))
            countp += 1
          elif i[12] == '1':
            treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('blue',))
            countp += 1
          else:
            pass
         
      sql = "select * from Productservice where Productserviceid = %s"
      val = (itemid, )
      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone()
      
      
      top = Toplevel()  
      top.title("Edit Product/Service details")
      p3 = PhotoImage(file = 'images/fbicon.png')
      top.iconphoto(False, p1)
      top.geometry("600x550+390+125")
      tabControl = ttk.Notebook(top)
      s = ttk.Style()
      s.theme_use('default')
      s.configure('TNotebook.Tab', background="#999999", width=50, padding=10,bd=0)

      taba = ttk.Frame(tabControl)
      tabb = ttk.Frame(tabControl)
      
      tabControl.add(taba,compound = LEFT, text ='Product/Service')
      tabControl.add(tabb,compound = LEFT, text ='Product Image')
      
      tabControl.pack(expand = 1, fill ="both")
      
      innerFrame = Frame(taba,bg="#f5f3f2", relief=GROOVE)
      innerFrame.pack(side="top",fill=BOTH)

      updateframe = LabelFrame(innerFrame,text="Product/Service",width=580,height=485)
      updateframe.pack(side="top",fill=BOTH,padx=10)

      code1=Label(updateframe,text="Code or SKU:",fg="blue",pady=10,padx=10)
      code1.place(x=20,y=0)
      codeentry = Entry(updateframe,width=35)
      codeentry.place(x=110,y=8)
      codeentry.insert(0, psdata[2])

      checkvarStatus=IntVar()
      status1=Label(updateframe,text="Status:")
      status1.place(x=380,y=8)
      Button1 = Checkbutton(updateframe, 
                        variable = checkvarStatus,text="Active",compound="right",
                        onvalue =1,
                        offvalue =0,
                        width = 10)
      Button1.place(x=420,y=5)
      sta = psdata[6]
      if sta == '1':
        Button1.select()
      else:
        Button1.deselect()



      category1=Label(updateframe,text="Category:",pady=5,padx=10)
      category1.place(x=20,y=40)
      n = StringVar() 
      category = Entry(updateframe,width=70,textvariable=n) 
      category.place(x=110,y=45)
      category.insert(0, psdata[3])


      name1=Label(updateframe,text="Name :",fg="blue",pady=5,padx=10)
      name1.place(x=20,y=70)
      nameentry = Entry(updateframe,width=70)
      nameentry.place(x=110,y=75)
      nameentry.insert(0, psdata[4])

      des1=Label(updateframe,text="Description :",pady=5,padx=10)
      des1.place(x=20,y=100)
      desentry = Entry(updateframe,width=70)
      desentry.place(x=110,y=105)
      desentry.insert(0, psdata[5])

      def set_label(name, index, mode):
        priceval.set(uval.get() - costval.get())
      
      unit1=Label(updateframe,text="Unit Price:",fg="blue",pady=5,padx=10)
      unit1.place(x=20,y=130)
      
      uval = IntVar()
      unitentry = Entry(updateframe,width=20,textvariable=uval)
      unitentry.place(x=110,y=135)
      unitentry.delete(0,'end')
      unitentry.insert(0, psdata[7])
      

      pcsval = IntVar()
      pcs1=Label(updateframe,text="Pcs/Weight:",fg="blue",pady=5,padx=10)
      pcs1.place(x=320,y=130)
      pcsentry = Entry(updateframe,width=20,textvariable=pcsval)
      pcsentry.place(x=410,y=135)
      pcsentry.delete(0,'end')
      pcsentry.insert(0, psdata[8])
      

      costval = IntVar()
      cost1=Label(updateframe,text="Cost:",pady=5,padx=10)
      cost1.place(x=20,y=160)
      costentry = Entry(updateframe,width=20,textvariable=costval)
      costentry.place(x=110,y=165)
      costentry.delete(0,'end')
      costentry.insert(0, psdata[9])
      

      priceval = IntVar()
      price1=Label(updateframe,text="(Price-Cost):",pady=5,padx=10)
      price1.place(x=20,y=190)
      priceentry = Entry(updateframe,width=20,textvariable=priceval)
      priceentry.place(x=110,y=195)
      priceentry.delete(0,'end')
      priceentry.insert(0, psdata[11])

      uval.trace('w', set_label)
      costval.trace('w', set_label)
      

      checkvarStatus2=IntVar()
    
      Button2 = Checkbutton(updateframe,variable = checkvarStatus2, 
                        text="Taxable Tax1rate",compound="right",
                        onvalue =1 ,
                        offvalue =0,
                        height=2,
                        width = 12)

      Button2.place(x=415,y=153)
      tax = psdata[10]
      if tax == '1':
        Button2.select()
      else:
        Button2.deselect()


      checkvarStatus3=IntVar()
    
      Button3 = Checkbutton(updateframe,variable = checkvarStatus3, 
                        text="No stock Control", 
                        onvalue =1 ,
                        offvalue = 0,
                        height=3,
                        width = 15)

      Button3.place(x=40,y=220)

      


      stockval = IntVar(updateframe)
      stock1=Label(updateframe,text="Stock:",pady=5,padx=10)
      stock1.place(x=90,y=260)
      stockentry = Entry(updateframe,width=15,textvariable=stockval)
      stockentry.place(x=140,y=265)
      stockentry.delete(0,'end')
      stockentry.insert(0, psdata[13])
      

      lowval = IntVar(updateframe)
      low1=Label(updateframe,text="Low Stock Warning Limit:",pady=5,padx=10)
      low1.place(x=280,y=260)
      lowentry = Entry(updateframe,width=15,textvariable=lowval)
      lowentry.place(x=435,y=265)
      lowentry.delete(0,'end')
      lowentry.insert(0, psdata[14])
      

    
      ware1=Label(updateframe,text="Warehouse:",pady=5,padx=10)
      ware1.place(x=60,y=290)
      wareentry = Entry(updateframe,width=64)
      wareentry.place(x=140,y=295)
      wareentry.insert(0, psdata[15])

      scr = psdata[12]
      if scr == '1':
        Button3.select()
        stockentry["state"] = DISABLED
        lowentry["state"] = DISABLED
        wareentry["state"] = DISABLED
      else:
        Button3.select()
        stockentry["state"] = NORMAL
        lowentry["state"] = NORMAL
        wareentry["state"] = NORMAL
      
      

      

      text1=Label(updateframe,text="Private notes(not appears on invoice):",pady=5,padx=10)
      text1.place(x=20,y=320)
      sctxt = scrolledtext.ScrolledText(updateframe, undo=True,width=62,height=4)
      sctxt.place(x=32,y=358)
      try:
        sctxt.insert("1.0", psdata[16])
      except:
        pass

      okButton = Button(innerFrame, text ="Ok",image=tick,width=70,compound = LEFT, command=updateproducts)
      okButton.pack(side=LEFT, padx=(10, 0))

      cancelButton = Button(innerFrame,image=cancel,text="Cancel",width=70,compound = LEFT, command=lambda : top.destroy())
      cancelButton.pack(side=RIGHT, padx=(0, 10))
      
      
      imageFrame = Frame(tabb, relief=GROOVE,height=580)
      imageFrame.pack(side="top",fill=BOTH)

      browseimg=Label(imageFrame,text=" Browse for product image file(recommended image type:JPG,size 480x320 pixels) ",bg='#f5f3f2')
      browseimg.place(x=15,y=35)

      browsebutton=Button(imageFrame,text = 'Browse', command=upload_file)
      browsebutton.place(x=470,y=30,height=30,width=50)

      try:
        image = Image.open("images/"+psdata[17])
        resize_image = image.resize((350, 350))
        image = ImageTk.PhotoImage(resize_image)
        b2 = Label(imageFrame,image=image,width=350,height=350)
        b2.photo = image
        b2.place(x=130, y=80)
        print(image)
      except:
        pass

      removeButton = Button(imageFrame,image=cancel,text="Remove Product Image",width=150,compound = LEFT)
      removeButton.place(x=410,y=460)
    except:
      try:
        top.destroy()
      except:
        pass
      messagebox.showerror('F-Billing Revolution', 'Select a record to edit.')
  


######################## DELETE PRODUCT #######################################################################


def delete_product():
    delmess = messagebox.askyesno("Delete product/service", "Are you sure to delete this product/service?")
    if delmess == True:
      itemid = treeproducts.item(treeproducts.focus())["values"][1]
      sql = "delete from Productservice where Productserviceid = %s"
      val = (itemid, )
      fbcursor.execute(sql, val)
      fbilldb.commit()
      treeproducts.delete(treeproducts.selection()[0])
      # messagebox.showinfo("F-Billing Revolution", "Record deleted successfully.")
    else:
      pass


######################## SEARCH PRODUCT  #######################################################################

def search_pro():
  query = searchvar.get()
  selections = []
  for child in treeproducts.get_children():
      if query in treeproducts.item(child)['values']:
          print(treeproducts.item(child)['values'])
          selections.append(child)
  treeproducts.selection_set(selections)
  

def search_product():
    global searchvar, searchtop
    searchtop = Toplevel()  
    searchtop.title("Find Text")
    searchtop.geometry("520x200+390+250")
    
    findwhat1=Label(searchtop,text="Find What:",pady=5,padx=10)
    findwhat1.place(x=5,y=20)
    searchvar = StringVar() 
    findwhat = ttk.Combobox(searchtop, width = 50, textvariable = searchvar)
    findwhat.place(x=80,y=25) 
    

    findButton = Button(searchtop, text ="Find next",width=10, command=search_pro)
    findButton.place(x=420,y=20)
    
    findin1=Label(searchtop,text="Find in:",pady=5,padx=10)
    findin1.place(x=5,y=47)
    n = StringVar() 
    findIN = ttk.Combobox(searchtop, width = 37, textvariable = n ) 
    # Adding combobox drop down list 
    findIN['values'] = ('Product/Service id',  
                              'Category', 
                              'Active', 
                              'name', 
                              'stock', 
                              'location', 
                              'image', 
                              '<<All>>') 
      
    findIN.place(x=80,y=54) 
    findIN.current(0)

    closeButton = Button(searchtop, text ="Close",width=10, command=lambda : searchtop.destroy())
    closeButton.place(x=420,y=50)

    match1=Label(searchtop,text="Match:",pady=5,padx=10)
    match1.place(x=5,y=74)
    n = StringVar() 
    match = ttk.Combobox(searchtop, width = 27, textvariable = n ) 
      
    # Adding combobox drop down list 
    match['values'] = ('From Any part',' Whole Field',  
                              ' From the beginning of the field')
      
    match.place(x=80,y=83) 
    match.current(0)

    search1=Label(searchtop,text="Search:",pady=5,padx=10)
    search1.place(x=5,y=102)
    n = StringVar() 
    search = ttk.Combobox(searchtop, width = 27, textvariable = n ) 
      
    # Adding combobox drop down list 
    search['values'] = ('All', 'up', 
                              'Down')
      
    search.place(x=80,y=112) 
    search.current(0)


    checkvarStatus4=IntVar()
   
    Button4 = Checkbutton(searchtop,variable = checkvarStatus4, 
                      text="Match Case", 
                      onvalue =0 ,
                      offvalue = 1,
                      height=3,
                      width = 15)

    Button4.place(x=60,y=141)

    checkvarStatus5=IntVar()
   
    Button5 = Checkbutton(searchtop,variable = checkvarStatus5, 
                      text="Match Format", 
                      onvalue =0 ,
                      offvalue = 1,
                      height=3,
                      width = 15)

    Button5.place(x=270,y=141)

    searchtop.mainloop()


######################## REFRESH PRODUCT  #######################################################################

def refresh_pro_s():
  for record in treeproducts.get_children():
    treeproducts.delete(record)
  fbcursor.execute("select *  from Productservice")
  pandsdata = fbcursor.fetchall()
  countp = 0
  for i in pandsdata:
    if i[6] == '1':
      acti = 'Active'
    else:
      acti = 'Inactive' 
    if i[13] > i[14]:
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
      countp += 1
    elif (i[12] =="0") == (i[13] <= i[14]):
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
      countp += 1
    elif i[12] == '1':
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
      countp += 1
    else:
      pass

######################## View Image
# #######################################################################
def psfile_image(event):
      itemid = treeproducts.item(treeproducts.focus())["values"][1]
      sql = "select * from Productservice where Productserviceid = %s"
      val = (itemid, )

      fbcursor.execute(sql, val)
      psdata = fbcursor.fetchone() 
      if psdata[17] is None:
        pass
      else:
        edit_window_img = Toplevel()
        edit_window_img.title("View Image")
        edit_window_img.geometry("700x500")
        image = Image.open("images/"+psdata[17])
        resize_image = image.resize((700, 500))
        image = ImageTk.PhotoImage(resize_image)
        psimage = Label(edit_window_img,image=image)
        psimage.photo = image
        psimage.pack()
  
######################## FRONT PAGE OF PRODUCT SERVICE SECTION #######################################################################

    
mainFrame=Frame(tab8, relief=GROOVE, bg="#f8f8f2")
mainFrame.pack(side="top", fill=BOTH)

midFrame=Frame(mainFrame, bg="#f5f3f2", height=60)
midFrame.pack(side="top", fill=X)

lFrame=Frame(tab8, bg="#f8f8f2", height=600)
lFrame.pack(side="top", fill=X)



pn = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=(5, 2))
pn = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=(0, 5))

productIcon = ImageTk.PhotoImage(Image.open("images/plus.png"))
productLabel = Button(midFrame,compound="top", text="Add new\nProduct",relief=RAISED,command=adda_product, image=productIcon, font=("arial", 8),bg="#f5f3f2", fg="black", height=55, bd=1, width=55)
productLabel.pack(side="left", pady=3, ipadx=4)

proeditIcon = ImageTk.PhotoImage(Image.open("images/edit.png"))
proeditLabel = Button(midFrame,compound="top", text="Edit/View\nProduct",relief=RAISED, image=proeditIcon,command=edit_product, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
proeditLabel.pack(side="left")

prodeleteIcon = ImageTk.PhotoImage(Image.open("images/delete.png"))
prodeleteLabel = Button(midFrame,compound="top", text="Delete\nSelected",relief=RAISED, command=delete_product,image=prodeleteIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
prodeleteLabel.pack(side="left")

pn = Canvas(midFrame, width=1, height=65, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=5)

prosearchIcon = ImageTk.PhotoImage(Image.open("images/research.png"))
prosearchLabel = Button(midFrame,compound="top", text="Search in\nproducts",relief=RAISED,command=search_product, image=prosearchIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55, activebackground="red")
prosearchLabel.pack(side="left")

proimportIcon = ImageTk.PhotoImage(Image.open("images/import.png"))
proimportLabel = Button(midFrame,compound="top", text="Import\nProducts",relief=RAISED,command=fileimport_product, image=proimportIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
proimportLabel.pack(side="left")

pn = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
pn.pack(side="left", padx=5)

proexportIcon = ImageTk.PhotoImage(Image.open("images/export-file.png"))
proexportLabel = Button(midFrame,compound="top",command=export_product, text="Export\nProducts",relief=RAISED, image=proexportIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55)
proexportLabel.pack(side="left")


pn = Canvas(midFrame, width=1, height=55, bg="#b3b3b3", bd=0)
pn.pack(side="left")

prorefreshIcon = ImageTk.PhotoImage(Image.open("images/refresh.png"))
productrefreshLabel = Button(midFrame,compound="top", text="Refresh\nProduct List",relief=RAISED, image=prorefreshIcon, font=("arial", 8),bg="#f8f8f2", fg="black", height=55, bd=1, width=55, command=refresh_pro_s)
productrefreshLabel.pack(side="left")

prolabel = Label(mainFrame, text="Products/Services", font=("arial", 18), bg="#f8f8f2")
prolabel.pack(side="left", padx=(20,0))

pr_label = Label(mainFrame, text="Category", font=("arial", 16), bg="#f8f8f2")
pr_label.place(x=1099,y=70)

def pro_fil(event):
  pro_f = dro.get()
  for record in treeproducts.get_children():
    treeproducts.delete(record)

  sql = "select * from Productservice where category = %s"
  val = (pro_f,)
  fbcursor.execute(sql, val)
  product_fil = fbcursor.fetchall()
  
  countp = 0
  for i in product_fil:
    if i[6] == '1':
      acti = 'Active'
    else:
      acti = 'Inactive' 
    if i[13] > i[14]:
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('green',))
      countp += 1
    elif (i[12] =="0") == (i[13] <= i[14]):
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('red',))
      countp += 1
    elif i[12] == '1':
      treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4] , i[7], i[13], i[15],i[17]),tags=('blue',))
      countp += 1
    else:
      pass

sql = "SELECT DISTINCT category FROM Productservice"
fbcursor.execute(sql,)
lic = fbcursor.fetchall()
dro = ttk.Combobox(mainFrame,)
dro.pack(side="right", padx=(0,10))
dro['values'] = lic
dro.bind("<<ComboboxSelected>>", pro_fil)

pro_label = Label(mainFrame, text="Right click on datagrid row for more options.",  bg="#f8f8f2")
pro_label.pack(side="right", padx=(0,260))


sql = 'select * from Productservice'
fbcursor.execute(sql)
pandsdata = fbcursor.fetchall()

treeproducts=ttk.Treeview(tab8,selectmode='browse')
treeproducts.place(x=8,y=100,height=580)
vertical_bar=ttk.Scrollbar(tab8,orient="vertical")
vertical_bar.place(x=1083,y=101,height=580)

treeproducts["columns"]=("1","2","3","4","5","6","7","8","9")
treeproducts["show"]='headings'
treeproducts.column("1",width=0,anchor='c', stretch=False)
treeproducts.column("2",width=160,anchor='c')
treeproducts.column("3",width=190,anchor='c')
treeproducts.column("4",width=120,anchor='c')
treeproducts.column("5",width=120,anchor='c')
treeproducts.column("6",width=120,anchor='c')
treeproducts.column("7",width=130,anchor='c')
treeproducts.column("8",width=120,anchor='c')
treeproducts.column("9",width=112,anchor='c')
treeproducts.heading("1",text="")
treeproducts.heading("2",text="Product/Service ID")
treeproducts.heading("3",text="Category")
treeproducts.heading("4",text="Status")
treeproducts.heading("5",text="Name")
treeproducts.heading("6",text="Price")
treeproducts.heading("7",text="Stock")
treeproducts.heading("8",text="Location/warehouse")
treeproducts.heading("9",text="Image")
treeproducts.bind('<Double-Button-1>' , psfile_image)
countp = 0
treeproducts.tag_configure('green', foreground='green')
treeproducts.tag_configure('red', foreground='red')
treeproducts.tag_configure('blue', foreground='blue')
for i in pandsdata:
  if i[6] == '1':
    acti = 'Active'
  else:
    acti = 'Inactive' 
  if i[13] > i[14]:
    treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
    countp += 1
  elif (i[12] =="0") == (i[13] <= i[14]):
    treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
    countp += 1
  elif i[12] == '1':
    treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
    countp += 1
  else:
    pass



treeproducts.place(height=580, width=1070, x=10, y=101)

######side_Listbox##############

treeps=ttk.Treeview(tab8,selectmode='browse')
treeps.place(height=580,width=254,
                      x=1099,y=101
                      )
treeps["columns"]=("1")
treeps["show"]='headings'
treeps.column("1",width=254,anchor='c')
treeps.heading("1",text="View filter by category")

def items_selected(event):
  selected_indices = listbox.curselection()
  selected_filter = ",".join([listbox.get(i) for i in selected_indices])

  sql = 'select * from Productservice'
  fbcursor.execute(sql)
  pandsdata = fbcursor.fetchall()
  psql = "select * from Productservice where category=%s"
  val = ('product', )
  fbcursor.execute(psql, val)
  pdata = fbcursor.fetchall()

  ssql = "select * from Productservice where serviceornot=%s"
  val = ('1', )
  fbcursor.execute(ssql, val)
  sdata = fbcursor.fetchall()

  # pssql = "select * from Productservice where category=%s"
  # psval = (selected_filter, )
  # fbcursor.execute(pssql, psval)
  # pssdata = fbcursor.fetchall()
  if selected_filter == "View all records":
    for record in treeproducts.get_children():
      treeproducts.delete(record)
    countp = 0
    for i in pandsdata:
      if i[6] == '1':
        acti = 'Active'
      else:
        acti = 'Inactive' 
      if i[13] > i[14]:
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
        countp += 1
      elif (i[12] =="0") == (i[13] <= i[14]):
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
        countp += 1
      elif i[12] == '1':
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
        countp += 1
      else:
        pass
  elif selected_filter == "View all products":
    for record in treeproducts.get_children():
      treeproducts.delete(record)
    countp = 0
    for i in pdata:
      if i[6] == '1':
        acti = 'Active'
      else:
        acti = 'Inactive' 
      if i[13] > i[14]:
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
        countp += 1
      elif (i[12] =="0") == (i[13] <= i[14]):
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
        countp += 1
      elif i[12] == '1':
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
        countp += 1
      else:
        pass
  elif selected_filter == "View all services":
    for record in treeproducts.get_children():
      treeproducts.delete(record)
    countp = 0
    for i in sdata:
      if i[6] == '1':
        acti = 'Active'
      else:
        acti = 'Inactive' 
      if i[13] > i[14]:
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('green',))
        countp += 1
      elif (i[12] =="0") == (i[13] <= i[14]):
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('red',))
        countp += 1
      elif i[12] == '1':
        treeproducts.insert(parent='', index='end', iid=countp, text='hello', values=('', i[0], i[3], acti, i[4], i[7], i[13], i[15],i[17]),tags=('blue',))
        countp += 1
      else:
        pass
  # elif selected_filter == True:
  #     for record in treeproducts.get_children():
  #       treeproducts.delete(record)
  #     countp = 0
  #     for i in pssdata:
  #       treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i    [2], i[4], i[7], i[13], i[15]))
  #       countp += 1
# def cat_pro(event):
#   selected_indices = listbox.curselection()
#   selected_filt = ",".join([listbox.get(i) for i in selected_indices])
#   pssql = "select * from Productservice where category=%s"
#   psval = (selected_filt, )
#   fbcursor.execute(pssql, psval)
#   pssdata = fbcursor.fetchall()
#   print(pssdata)
#   for record in treeproducts.get_children():
#     treeproducts.delete(record)
#   countp = 0
#   for i in pssdata:
#     treeproducts.insert(parent='', index='end', iid=countp, text='servicedata', values=('', i[0], i[3], i[2], i[4], i[7], i[13], i[15]))
#     countp += 1


# sql = "SELECT DISTINCT category FROM Productservice WHERE NOT category = %s AND NOT category = %s"
# val = ('service','product',)
# fbcursor.execute(sql,val,)
# lic = fbcursor.fetchall()
# print(lic)
listbox = Listbox(tab8,height = 8,  
                      width = 29,  
                      bg = "white",
                      activestyle = 'dotbox',  
                      fg = "black",
                      highlightbackground="white")  
listbox.insert(0, "View all records")
listbox.insert(1, "View all products")
listbox.insert(2, "View all services")
# for nc in lic:
#     listbox.insert(END, nc)

listbox.place(x=1099,y=118,height=564,width=255)

listbox.bind('<<ListboxSelect>>', items_selected)
# listbox.bind('<<ListboxSelect-1>>', cat_pro)

stockok = Label(tab8,text="Green: Stock is Ok",foreground="green",background="white").place(x =1110,y =580)
stocko = Label(tab8,text="Red: Limit <= Low Stock Limit",foreground="red",background="white").place(x =1110,y =600)
stock = Label(tab8,text="Blue: Service,no Stock Control",foreground="blue",background="white").place(x =1110,y =620)




root.mainloop()